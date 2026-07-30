"""Phase 9 - assemble out/rfp_pages.csv and run the mandatory sanity checks.

Input columns are copied from out/agencies.json, which stored them verbatim from
the XLSX, so they are byte-identical to the source. Hosting/platform are derived
here (not at discovery time) so the corrected A2 domain rule applies uniformly.
"""
from __future__ import annotations

import csv
import json
import re
import sys
from urllib.parse import urlparse

import common as C

COLS = ["agency_name", "state", "agency_type", "agency_website",
        "rfp_url", "rfp_hosting", "rfp_platform", "notes"]

HOSTING_ENUM = {"self-hosted", "third-party", "not-found"}

SIGNAL_WORDS = {
    "sitemap": "found via sitemap.xml",
    "path-probe": "found by probing canonical paths",
    "site-search": "found via the site's search endpoint",
    "home-link": "found via a homepage link",
    "depth2-link": "found via a second-level page link",
    "portal-link": "portal linked from the agency homepage",
    "portal-via-page": "portal linked from the agency's own bids page",
    "taxonomy-archive": "found via the site's category archive",
    "cms:egov-doccenter": "eGov document centre, 'Bids' document type",
    "rendered-link": "found via a JS-rendered nav link",
    "rendered-probe": "found by probing canonical paths in a real browser",
    "rendered-hop2": "found one hop from a rendered candidate page",
    "site-index": "found via the site's A-Z / index page",
    "tier3-agent": "resolved by agent navigation",
    "a9-agent": "resolved by agent navigation (no website in input)",
}


def blocked_evidence():
    """Scan EVERY checkpoint line for access failures.

    The merged 'best record' per agency can come from a tier that saw no block,
    which silently dropped the 403 evidence from Spencer's note. A block is the
    reason a row is empty and must survive into the deliverable (A6).
    """
    out = {}
    path = C.CHECKPOINT
    try:
        fh = open(path, encoding="utf-8")
    except FileNotFoundError:
        return out
    with fh as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                r = json.loads(line)
            except Exception:
                continue
            k = str(r.get("idx"))
            hint = r.get("notes_hint") or ""
            if r.get("browser_blocked"):
                out[k] = hint or "blocked: confirmed in a real browser (A6)"
            elif "blocked" in hint and k not in out:
                out[k] = hint
    return out


def build_note(rec, agency, hosting, platform, extra):
    bits = []
    sig = rec.get("signal", "")
    if rec.get("a9_inferred"):
        bits.append(f"website absent from input; domain inferred as "
                    f"{rec['a9_inferred']} by name+Indiana self-identification "
                    f"- VERIFY MANUALLY (A9)")
    if rec.get("status") == "resolved":
        bits.append(SIGNAL_WORDS.get(sig, f"found via {sig}" if sig else "found"))
        if rec.get("via"):
            bits.append(f"via {rec['via']}")
        if rec.get("empty_state"):
            bits.append("no open solicitations listed at capture time (page is correct, A8)")
        elif rec.get("items"):
            bits.append(f"{rec['items']} solicitation item(s) on page")
        elif not rec.get("tier3_note"):
            # Be explicit when the page matched on path/title but no individual
            # solicitation could be parsed - the reviewer should know which is which.
            # (Skipped when an agent read the page and said what is on it.)
            bits.append("page identified by URL+title; no individual solicitation "
                        "entries could be parsed (may be JS- or PDF-published)")
        if extra and not extra.startswith("CMS:"):
            bits.append(extra)
        if rec.get("cms") and hosting == C.HOSTING_SELF:
            cms = re.sub(r"\s*\(.*\)\s*$", "", rec["cms"]).strip()
            bits.append(f"CMS: {cms}")
        if rec.get("tier3_note"):
            bits.append(rec["tier3_note"])
        if rec.get("recheck"):
            bits.append(rec["recheck"])
        if rec.get("score") is not None:
            bits.append(f"verify score {rec['score']}")
    else:
        reason = (rec.get("not_found_reason")
                  or "no RFP/bid listing page found on the agency's own site")
        bits.append(reason)
        hint = rec.get("notes_hint") or rec.get("_blocked") or ""
        # don't repeat a block/mismatch already spelled out in the reason
        if hint and hint.split(":")[0] not in reason:
            bits.append(hint)
        elif rec.get("_blocked") and "blocked" not in reason and hint != rec["_blocked"]:
            bits.append(rec["_blocked"])
    # de-duplicate while preserving order (cms was being appended twice)
    seen, out = set(), []
    for b in bits:
        b = (b or "").strip().rstrip(";")
        if b and b.lower() not in seen:
            seen.add(b.lower())
            out.append(b)
    return "; ".join(out)


def main():
    agencies = C.load_agencies()
    done = C.checkpoint_read()
    # out/final.jsonl is authoritative when present: it is the adjudication pass,
    # the only stage that can DEMOTE a URL an earlier tier accepted.
    try:
        with open(C.OUT + "/final.jsonl", encoding="utf-8") as f:
            final = {str(json.loads(l)["idx"]): json.loads(l)
                     for l in f if l.strip()}
        print(f"using out/final.jsonl (adjudicated) for {len(final)} agencies")
        for k, v in final.items():
            merged = dict(done.get(k) or {})
            merged.update(v)
            done[k] = merged
    except FileNotFoundError:
        print("no out/final.jsonl - falling back to raw checkpoints")
    # Tier 3 (agent read the page) is applied last and wins outright.
    try:
        with open(C.OUT + "/tier3_decisions.json", encoding="utf-8") as f:
            t3 = json.load(f)
        print(f"applying {len(t3)} Tier-3 agent decisions")
        for k, v in t3.items():
            merged = dict(done.get(k) or {})
            merged.update(v)
            done[k] = merged
    except FileNotFoundError:
        pass
    rows, problems = [], []
    blocked = blocked_evidence()
    try:
        with open(C.OUT + "/a9_inferred.json", encoding="utf-8") as f:
            a9inf = json.load(f)
    except FileNotFoundError:
        a9inf = {}

    for a in agencies:
        rec = done.get(str(a["idx"])) or {}
        url = rec.get("rfp_url") or ""
        if rec.get("status") != "resolved":
            url = ""
        if blocked.get(str(a["idx"])) and not url:
            rec = dict(rec)
            rec["_blocked"] = blocked[str(a["idx"])]
        inf = a9inf.get(str(a["idx"]))
        if inf:
            # A9: the input has no domain, so hosting is judged against the
            # inferred host, and the row carries a manual-verification flag.
            rec = dict(rec)
            rec["a9_inferred"] = inf["inferred_website"]
        if url:
            hosting, platform, extra = C.classify(
                url,
                (inf or {}).get("inferred_domain", a["agency_domain"]),
                (inf or {}).get("inferred_website", a["agency_website"]),
                rec.get("cms", ""))
        else:
            hosting, platform, extra = C.HOSTING_NONE, C.HOSTING_NONE, ""
        rows.append({
            "agency_name": a["agency_name"],
            "state": a["state"],
            "agency_type": a["agency_type"],
            "agency_website": a["agency_website"],
            "rfp_url": url,
            "rfp_hosting": hosting,
            "rfp_platform": platform,
            "notes": build_note(rec, a, hosting, platform, extra),
        })

    with open(C.FINAL_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLS)
        w.writeheader()
        w.writerows(rows)

    # ---------------- validation ----------------
    print(f"wrote {C.FINAL_CSV}")
    if len(rows) != 200:
        problems.append(f"row count is {len(rows)}, expected 200")

    # input columns byte-identical to the source xlsx
    from openpyxl import load_workbook
    wb = load_workbook(C.INPUT_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    src = [list(r) for r in ws.iter_rows(values_only=True)][1:]
    wb.close()
    src = [r for r in src if any(c not in (None, "") for c in r)]
    mismatch = 0
    for i, (r, s) in enumerate(zip(rows, src)):
        cells = list(s) + [None] * (5 - len(s))
        for j, col in enumerate(["agency_name", "state", "agency_type", "agency_website"]):
            want = "" if cells[j] is None else str(cells[j]).strip()
            if r[col] != want:
                mismatch += 1
                if mismatch <= 5:
                    problems.append(f"row {i+2} {col}: {r[col]!r} != {want!r}")
    if mismatch:
        problems.append(f"{mismatch} input-column mismatches")

    bad_host = {r["rfp_hosting"] for r in rows} - HOSTING_ENUM
    if bad_host:
        problems.append(f"invalid rfp_hosting values: {bad_host}")
    for r in rows:
        if not r["rfp_url"] and r["rfp_platform"] != "not-found":
            problems.append(f"{r['agency_name']}: blank url but platform {r['rfp_platform']}")
        if not r["rfp_url"] and not r["notes"].strip():
            problems.append(f"{r['agency_name']}: blank rfp_url with no note")
        if r["rfp_url"] and r["rfp_hosting"] == "not-found":
            problems.append(f"{r['agency_name']}: url present but hosting not-found")

    seen = {}
    for r in rows:
        if not r["rfp_url"]:
            continue
        k = C.norm_url(r["rfp_url"])
        if k in seen:
            problems.append(f"duplicate rfp_url: {r['agency_name']} == {seen[k]} -> {r['rfp_url']}")
        seen[k] = r["agency_name"]

    filled = sum(1 for r in rows if r["rfp_url"])
    import collections
    print(f"rows            : {len(rows)}")
    print(f"rfp_url filled  : {filled}   blank: {len(rows)-filled}")
    print(f"hosting         : {dict(collections.Counter(r['rfp_hosting'] for r in rows))}")
    print("platforms       :", dict(collections.Counter(
        r["rfp_platform"] for r in rows if r["rfp_platform"] not in ("self-hosted", "not-found"))))
    print(f"by type filled  : " + str({
        t: sum(1 for r in rows if r["agency_type"] == t and r["rfp_url"])
        for t in sorted({r["agency_type"] for r in rows})}))

    if problems:
        print(f"\n!! {len(problems)} VALIDATION PROBLEM(S):")
        for p in problems[:25]:
            print("   -", p)
        sys.exit(1)
    print("\nAll validation checks passed.")


if __name__ == "__main__":
    main()
