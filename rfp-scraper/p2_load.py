"""Phase 2 - loader.

Reads the XLSX verbatim, normalizes short rows to 5 cells, splits the A9 lane
(agencies with no website/domain), writes out/agencies.json.
Values are stored EXACTLY as read so phase 9 can emit byte-identical columns.
"""
from __future__ import annotations

import json

from openpyxl import load_workbook

import common as C

EXPECTED = ["agency_name", "state", "agency_type", "agency_website", "agency_domain"]


def main():
    wb = load_workbook(C.INPUT_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    header = [(c if c is not None else "") for c in rows[0]]
    header = [str(h).strip() for h in header]
    body = rows[1:]
    # drop fully-empty trailing rows
    body = [r for r in body if any(c not in (None, "") for c in r)]

    print(f"file       : {C.INPUT_XLSX}")
    print(f"rows       : {len(rows)} (1 header + {len(body)} agencies)")
    print(f"header     : {header}")
    assert header[:5] == EXPECTED, f"unexpected header: {header}"
    assert len(body) == 200, f"expected 200 agencies, got {len(body)}"

    agencies = []
    for i, r in enumerate(body, start=1):
        cells = list(r) + [None] * (5 - len(r))          # normalize 3-cell -> 5
        name, state, atype, site, dom = [
            ("" if c is None else str(c).strip()) for c in cells[:5]]
        anchored = bool(site or dom)
        agencies.append({
            "idx": i,
            "row": i + 1,                                 # 1-based xlsx row
            "agency_name": name,
            "state": state,
            "agency_type": atype,
            "agency_website": site,                       # verbatim, incl. trailing slash
            "agency_domain": dom,
            "lane": "main" if anchored else "A9",
        })

    types = {}
    for a in agencies:
        types[a["agency_type"]] = types.get(a["agency_type"], 0) + 1
    a9 = [a for a in agencies if a["lane"] == "A9"]
    main = [a for a in agencies if a["lane"] == "main"]

    # consistency checks from SETUP.md §7
    dom_in_site = sum(1 for a in main if a["agency_domain"] and
                      a["agency_domain"].lower() in a["agency_website"].lower())
    https = sum(1 for a in main if a["agency_website"].lower().startswith("https"))
    slash = sum(1 for a in main if a["agency_website"].endswith("/"))
    names = [a["agency_name"] for a in agencies]
    doms = [a["agency_domain"] for a in main]

    print(f"types      : {types}")
    print(f"states     : {sorted({a['state'] for a in agencies})}")
    print(f"anchored   : {len(main)}   A9 (no website/domain): {len(a9)}")
    print(f"domain in website: {dom_in_site}/{len(main)}")
    print(f"https      : {https}/{len(main)}   trailing slash: {slash}/{len(main)}")
    print(f"dup names  : {len(names) - len(set(names))}   dup domains: {len(doms) - len(set(doms))}")
    print("A9 rows    : " + "; ".join(f"{a['row']}:{a['agency_name']}" for a in a9))

    with open(C.AGENCIES_JSON, "w", encoding="utf-8") as f:
        json.dump(agencies, f, ensure_ascii=False, indent=1)
    print(f"wrote      : {C.AGENCIES_JSON}")


if __name__ == "__main__":
    main()
